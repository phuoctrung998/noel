"""
excel_extractor.py
==================
Trích xuất cấu trúc + công thức từ file Excel lớn để phân tích bằng Claude.

Cài đặt:
    pip install openpyxl

Sử dụng:
    python excel_extractor.py <đường_dẫn_file.xlsx>

Output:
    - <tên_file>_analysis.txt  →  paste toàn bộ vào Claude
    - <tên_file>_formulas.txt  →  chỉ công thức (nếu cần hỏi riêng)
"""

import sys
import json
import re
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Thiếu thư viện. Chạy: pip install openpyxl")
    sys.exit(1)


# ─── Cấu hình ──────────────────────────────────────────────────────────────────

MAX_HEADER_ROWS   = 5     # Số hàng đầu lấy làm header mẫu
MAX_SAMPLE_ROWS   = 3     # Số hàng dữ liệu mẫu (không phải công thức)
MAX_FORMULA_CELLS = 300   # Tối đa số ô công thức trích xuất mỗi sheet
MAX_COLS_PREVIEW  = 30    # Tối đa số cột hiển thị trong preview


# ─── Helpers ───────────────────────────────────────────────────────────────────

def cell_addr(row, col):
    return f"{get_column_letter(col)}{row}"


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


def detect_cross_sheet_refs(formula: str):
    """Tìm tất cả tham chiếu xuyên sheet trong công thức."""
    # Dạng: SheetName!A1 hoặc 'Sheet Name'!A1
    pattern = r"'?([^'!]+)'?!"
    return list(set(re.findall(pattern, formula)))


def classify_formula(formula: str):
    """Phân loại công thức theo nhóm chức năng."""
    f = formula.upper()
    if any(k in f for k in ["SUMIF", "SUMIFS", "SUMPRODUCT"]):
        return "Tổng có điều kiện"
    if any(k in f for k in ["SUM(", "+SUM"]):
        return "Tổng đơn giản"
    if any(k in f for k in ["VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "XLOOKUP"]):
        return "Tra cứu/Lookup"
    if any(k in f for k in ["IF(", "IFS(", "IFERROR", "IFNA"]):
        return "Điều kiện IF"
    if any(k in f for k in ["COUNTIF", "COUNTIFS", "COUNT("]):
        return "Đếm"
    if any(k in f for k in ["AVERAGE", "MAX(", "MIN(", "MEDIAN"]):
        return "Thống kê"
    if "!" in formula:
        return "Tham chiếu xuyên sheet"
    return "Khác"


# ─── Core extraction ───────────────────────────────────────────────────────────

def extract_sheet_info(ws, sheet_name: str, row_limit: int = None):
    """Trích xuất thông tin đầy đủ của một sheet."""
    # read_only mode đôi khi trả về max_row/max_column sai (=1) — dùng giá trị lớn
    # và để iter_rows tự dừng khi hết dữ liệu
    _max_row = ws.max_row or 1
    _max_col = ws.max_column or 1
    # Nếu openpyxl báo 1×1 nhưng file có thể lớn hơn, dùng fallback lớn
    if _max_row <= 1 and _max_col <= 1:
        _max_row = 100_000
        _max_col = 500
    max_row = _max_row if row_limit is None else min(_max_row, row_limit)
    max_col = _max_col

    info = {
        "name": sheet_name,
        "dimensions": f"{max_row} hàng × {max_col} cột",
        "max_row": max_row,
        "max_col": max_col,
        "headers": [],
        "sample_data": [],
        "formulas": [],           # list of dict
        "formula_stats": defaultdict(int),
        "cross_sheet_refs": set(),
        "named_ranges": [],
        "summary_rows": [],       # Hàng có tổng cộng / lũy kế
    }

    # ── Headers (hàng đầu tiên không trống) ──
    for row in ws.iter_rows(min_row=1, max_row=MAX_HEADER_ROWS, max_col=min(max_col, MAX_COLS_PREVIEW)):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in row]
        if any(row_vals):
            info["headers"].append(row_vals)

    # ── Dữ liệu mẫu & công thức ──
    sample_count = 0
    formula_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col), start=1):
        row_has_formula = False

        for col_idx, cell in enumerate(row, start=1):
            # read_only mode có thể trả về EmptyCell — lấy value an toàn
            try:
                val = cell.value
            except Exception:
                continue
            if val is None:
                continue

            if is_formula(val):
                row_has_formula = True
                formula_count += 1
                cat = classify_formula(val)
                info["formula_stats"][cat] += 1
                refs = detect_cross_sheet_refs(val)
                info["cross_sheet_refs"].update(refs)

                # Tính địa chỉ ô từ row_idx/col_idx thay vì dùng cell.row/cell.column
                addr = cell_addr(row_idx, col_idx)

                if formula_count <= MAX_FORMULA_CELLS:
                    info["formulas"].append({
                        "cell":      addr,
                        "formula":   val,
                        "category":  cat,
                        "cross_refs": refs,
                    })
            elif sample_count < MAX_SAMPLE_ROWS * max_col:
                if row_idx <= MAX_HEADER_ROWS + MAX_SAMPLE_ROWS:
                    sample_count += 1

        # Phát hiện hàng tổng/lũy kế — đọc an toàn tránh EmptyCell
        try:
            first_cell_val = str(row[0].value or "").lower()
        except Exception:
            first_cell_val = ""
        if any(k in first_cell_val for k in ["total", "tổng", "cộng", "lũy kế", "cumul", "grand"]):
            info["summary_rows"].append(row_idx)

    info["total_formulas"] = formula_count
    info["cross_sheet_refs"] = sorted(info["cross_sheet_refs"])
    return info


def extract_named_ranges(wb):
    """Lấy danh sách named ranges trong workbook."""
    named = []
    for name, defn in wb.defined_names.items():
        try:
            destinations = list(defn.destinations)
            sheets = [d[0] for d in destinations]
        except Exception:
            sheets = []
        named.append({
            "name":      name,
            "refers_to": str(defn.value),
            "sheets":    sheets,
        })
    return named


def build_dependency_map(all_sheets_info):
    """Xây dựng bản đồ phụ thuộc giữa các sheet."""
    dep_map = defaultdict(set)
    for info in all_sheets_info:
        sheet = info["name"]
        for f in info["formulas"]:
            for ref_sheet in f["cross_refs"]:
                if ref_sheet != sheet:
                    dep_map[sheet].add(ref_sheet)
    return {k: sorted(v) for k, v in dep_map.items()}


# ─── Report builder ────────────────────────────────────────────────────────────

def build_report(wb, file_path: Path, sheets_info, named_ranges, dep_map):
    lines = []
    sep  = "=" * 70
    sep2 = "-" * 50

    lines += [
        sep,
        f"BÁO CÁO PHÂN TÍCH CẤU TRÚC EXCEL",
        f"File: {file_path.name}",
        f"Tổng số sheet: {len(wb.sheetnames)}",
        f"Danh sách sheet: {', '.join(wb.sheetnames)}",
        sep, "",
    ]

    # ── Named Ranges ──
    if named_ranges:
        lines += ["## NAMED RANGES (Vùng đặt tên)", sep2]
        for nr in named_ranges:
            lines.append(f"  {nr['name']}  →  {nr['refers_to']}")
        lines.append("")

    # ── Dependency map ──
    if dep_map:
        lines += ["## BẢN ĐỒ PHỤ THUỘC GIỮA CÁC SHEET", sep2]
        for sheet, deps in dep_map.items():
            lines.append(f"  [{sheet}]  lấy dữ liệu từ: {', '.join(deps)}")
        lines.append("")

    # ── Chi tiết từng sheet ──
    for info in sheets_info:
        lines += [
            f"## SHEET: {info['name']}",
            sep2,
            f"Kích thước : {info['dimensions']}",
            f"Tổng công thức: {info['total_formulas']}",
        ]

        if info["cross_sheet_refs"]:
            lines.append(f"Tham chiếu tới sheet khác: {', '.join(info['cross_sheet_refs'])}")

        if info["summary_rows"]:
            lines.append(f"Hàng tổng/lũy kế phát hiện: {info['summary_rows']}")

        # Headers
        if info["headers"]:
            lines.append("\n### Cấu trúc cột (header):")
            for h_row in info["headers"]:
                lines.append("  | " + " | ".join(h_row[:MAX_COLS_PREVIEW]))

        # Formula stats
        if info["formula_stats"]:
            lines.append("\n### Thống kê loại công thức:")
            for cat, cnt in sorted(info["formula_stats"].items(), key=lambda x: -x[1]):
                lines.append(f"  {cat:<30} {cnt:>5} ô")

        # Formulas detail
        if info["formulas"]:
            lines.append(f"\n### Danh sách công thức (tối đa {MAX_FORMULA_CELLS} ô):")
            # Nhóm theo loại
            by_cat = defaultdict(list)
            for f in info["formulas"]:
                by_cat[f["category"]].append(f)

            for cat, flist in by_cat.items():
                lines.append(f"\n  [{cat}]")
                for f in flist:
                    cross = f"  ← sheet: {', '.join(f['cross_refs'])}" if f["cross_refs"] else ""
                    lines.append(f"    {f['cell']}: {f['formula']}{cross}")

        lines += ["", sep2, ""]

    return "\n".join(lines)


def build_formula_only_report(sheets_info):
    """Báo cáo chỉ chứa công thức — dùng để hỏi Claude về logic cụ thể."""
    lines = ["# TOÀN BỘ CÔNG THỨC THEO SHEET", ""]
    for info in sheets_info:
        if not info["formulas"]:
            continue
        lines.append(f"## Sheet: {info['name']}")
        for f in info["formulas"]:
            lines.append(f"  {f['cell']}: {f['formula']}")
        lines.append("")
    return "\n".join(lines)


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Cách dùng: python excel_extractor.py <file.xlsx>")
        sys.exit(1)

    file_path = Path(sys.argv[1])
    if not file_path.exists():
        print(f"❌ Không tìm thấy file: {file_path}")
        sys.exit(1)

    print(f"⏳ Đang đọc file: {file_path.name}  ({file_path.stat().st_size / 1_000_000:.1f} MB) ...")

    # Dùng read_only=True để đọc nhanh file lớn, data_only=False để giữ công thức
    wb = load_workbook(file_path, read_only=True, data_only=False)
    print(f"✅ Đọc xong. Sheets: {wb.sheetnames}")

    # Trích xuất từng sheet
    sheets_info = []
    for name in wb.sheetnames:
        print(f"   Đang xử lý sheet: {name} ...")
        ws = wb[name]
        info = extract_sheet_info(ws, name)
        sheets_info.append(info)
        print(f"     → {info['dimensions']}, {info['total_formulas']} công thức")

    named_ranges = extract_named_ranges(wb)
    dep_map      = build_dependency_map(sheets_info)

    # Xuất báo cáo chính
    report       = build_report(wb, file_path, sheets_info, named_ranges, dep_map)
    formula_rep  = build_formula_only_report(sheets_info)

    out_main    = file_path.with_name(file_path.stem + "_analysis.txt")
    out_formula = file_path.with_name(file_path.stem + "_formulas.txt")

    out_main.write_text(report, encoding="utf-8")
    out_formula.write_text(formula_rep, encoding="utf-8")

    # Tóm tắt
    total_formulas = sum(i["total_formulas"] for i in sheets_info)
    print("\n" + "=" * 50)
    print(f"✅ Hoàn thành!")
    print(f"   Tổng công thức đã quét: {total_formulas:,}")
    print(f"   Named ranges: {len(named_ranges)}")
    print(f"\n📄 File output:")
    print(f"   {out_main}       ← PASTE FILE NÀY VÀO CLAUDE")
    print(f"   {out_formula}    ← Chỉ công thức (hỏi riêng nếu cần)")
    print("\n💡 Gợi ý prompt cho Claude Opus 4.7:")
    print('   "Đây là cấu trúc file Excel thu mua của tôi.')
    print('    Hãy giải thích toàn bộ logic, đặc biệt các công thức lũy kế')
    print('    và quan hệ phụ thuộc giữa các sheet."')


if __name__ == "__main__":
    main()